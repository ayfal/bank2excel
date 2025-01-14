import sys, pathlib
from hebextract import get_hebrew_text, get_hebrew_text_chars
import tkinter as tk
from tkinter import filedialog, messagebox



#todo: the functions for getting the words should be in the hebextract.py file

# def get_next_word(text, str):
#     position = text.find(str)
#     if position == -1:
#         return "***not found***"
#     else:
#         position += len(str)
#         position = text.find(" ", position)
#         return text[position:text.find(" ", position+1)]


# def get_next_words(text, str, n):
#     next_words = get_next_word(text, str)
#     for i in range(n-1):
#         next_words += get_next_word(text, next_words)
#     return next_words

'''
this function gets a string and a list of words and a number n, and returns the n words after the string in the list
'''
def get_words_after(words, str, n):
    next_words = ""
    words_str=str.split()
    for i in range(len(words)-len(words_str)):
        str_found = True
        for j in range(len(words_str)):
            if words_str[j] not in words[i+j]:
                str_found = False
                break
        if str_found:
            for j in range(n):
                next_words += words[i+len(words_str)+j]+" "
    return next_words


'''
this function is meant for getting the boundries of a string
'''
def get_boundries(words, str):
    words_str=str.split()
    for p, page in enumerate(words):
        for i in range(len(page)):
            str_found = True
            for j in range(len(words_str)):
                if words_str[j] != page[i+j][4]:
                    str_found = False
                    break
            if str_found:
                return p, page[i][2], page[i+len(words_str)-1][0], page[i][1] # page_number, right_x, left_x, y
    return None, None, None, None



'''
this function is meant for getting columns from tables.
it gets a list of words with their coordinates, a string, and a number n, and returns the n words under the string in the list
'''
def get_words_under(words, str, n):
    next_words = []
    p, right_x, left_x, y = get_boundries(words, str)
    for i in range(len(words[p])):
        if words[p][i][2]<right_x+10 and words[p][i][0]>left_x-10 and words[p][i][1]>y:
                next_words.append(words[p][i])
    #sort next_words by y
    next_words.sort(key=lambda x: x[1])                
    #extract the words
    words_str = ""
    for i in range(n):
        words_str += next_words[i][4]+" "
    return words_str


def extract_data(pdf,xlsx,bank,cells):
    text, words_coordinates = get_hebrew_text(pdf)  # get text from document
    
    #values to extract: A: report_date, B: end_date, C: interest_type, D: current_interest, E: linkage_type F: balance, G: commissions, H: loan_sum, I: payment_method
    report_date = cells[0]
    end_date = cells[1]
    interest_type = cells[2]
    current_interest = cells[3]
    linkage_type = cells[4]
    balance = cells[5]
    commission = cells[6]
    loan_sum = cells[7]
    payment_method = cells[8]
    
    #now we'll insert the extracted data to excel
    from openpyxl import Workbook
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active


    if bank=='BenLeumi':
        #get loans
        loans=text.split(". מספר_הלוואה")
        for loan in loans:
            words=loan.split()
            if not "ת.סיום/סילוק" in loan:
                #get report_date:
                ws[report_date] = words[0]
                #skip the first item in the list, as it is not a loan:
                continue
            #get end_date:
            ws[end_date] = get_words_after(words, "ת.סיום/סילוק משנה בפועל", 1)
            #get interest_type:
            ws[interest_type] = get_words_after(words, "סוג ריבית", 3)
            #get current_interest:
            ws[current_interest] = get_words_after(words, "ריבית נוכחית ליום מסירת המידע", 2)
            #get linkage_type:
            ws[linkage_type] = get_words_after(words, "סוג הצמדה", 2)
            commissions = [0,0,0]
            #get commissions:
            try:
                commissions[0] = float(get_words_after(words, "עמלת אי מתן הודעה", 1))
            except:
                commissions[0] = 0
            try: 
                commissions[1] = float(get_words_after(words, "עמלת מדד ממוצע", 1))
            except:
                commissions[1] = 0
            try:
                commissions[2] = float(get_words_after(words, "עמלת היוון בגין הפרשי ריבית", 1))
            except:
                commissions[2] = 0
            ws[commission] = sum(commissions)
            #get loan_sum:
            ws[loan_sum] = get_words_after(words, 'סה"כ להלוואה', 1).split()[1]
            #get payment_method:
            ws[payment_method] = get_words_after(words, 'שיטת פרעון ההלוואה', 1)
            #get balance (H-G)
            loan_sum_clean = ws[loan_sum].value.replace(',', '')
            H_G = float(loan_sum_clean) - sum(commissions)
            ws[balance] = H_G


    if bank=='Discount':
        words=text.split()
        #get loans
        loans=text.split("מספר   הלוואה")
        for loan in loans:
            words=loan.split()       
            if not "מועד תשלום אחרון" in loan:
                #get report_date:
                ws[report_date] = get_words_after(words, "תאריך הדפסה :", 1)
                continue
            #get end_date
            ws[end_date] = get_words_after(words, "מועד תשלום אחרון", 1)
            #get interest_type
            ws[interest_type] = get_words_after(words, "סוג ריבית", 1)
            #get current_interest
            ws[current_interest] = get_words_after(words, "שיעור ריבית שנתית", 1)
            #get linkage_type
            ws[linkage_type] = get_words_after(words, "בסיס הצמדה", 1)
            #get balance
            ws[balance] = get_words_after(words, 'סה"כ יתרה', 1)
            #get commission
            ws[commission] = get_words_after(words, 'סה"כ עמלת פרעון מוקדם', 1)
            #get loan_sum
            ws[loan_sum] = get_words_after(words, 'סה"כ יתרה לסילוק', 1)
            #get payment_method
            ws[payment_method] = get_words_after(words, 'שיטת פרעון', 1)

    if bank=='Leumi':
        #get loans
        loans=text.split("מספר משנה")
        for loan in loans:
            words=loan.split()       
            if not "תאריך סיום" in loan:
                #get report_date:
                ws[report_date] = get_words_after(words, "ליום:", 1)
                continue
            #get end_date
            ws[end_date] = get_words_after(words, "תאריך סיום", 1)
            #get interest_type
            ws[interest_type] = get_words_after(words, "סוג ריבית", 1)
            #get current_interest
            ws[current_interest] = get_words_after(words, "ריבית שנתית להלוואה", 1)
            #get linkage_type
            ws[linkage_type] = get_words_after(words, "בסיס הצמדה", 1)
            #get balance
            ws[balance] = get_words_after(words, 'יתרת קרן', 1)
            #get commission
            ws[commission] = get_words_after(words, 'סה"כ עמלות פירעון מוקדם', 1)
            #get loan_sum
            ws[loan_sum] = get_words_after(words, 'סה"כ יתרה לסילוק', 1)
    
    if bank=='Poalim':
        words=text.split()
        #get A: report_date:
        ws[report_date] = get_words_after(words, "תאריך :", 1)
        #get B: end_date
        ws[end_date] = get_words_after(words, "מועד צפוי לתשלום", 3)
        #get C: interest_type
        ws[interest_type] = get_words_after(words, "סוג ריבית", 3)
        #get D: current_interest
        ws[current_interest] = get_words_after(words, "שיעור ריבית נומינלית", 3)
        #get E: linkage_type
        ws[linkage_type] = get_words_after(words, "בסיס ההצמדה", 9)
        #get G: commission
        ws[commission] = get_words_under(words_coordinates, 'עמלת פירעון מוקדם', 6)
        #get H: loan_sum
        ws[loan_sum] = get_words_under(words_coordinates, 'יתרה לסילוק', 6)
        #get I: payment_method
        ws[payment_method] = get_words_after(words, 'שיטת פירעון ההלוואה', 3)

    if bank=='Tfakhot':
        text, words_coordinates = get_hebrew_text_chars(pdf)  # get text from document
        words=text.split()
        #get A: report_date:
        ws[report_date] = get_words_after(words, "יום עסקים", 1)
        #get B: end_date
        ws[end_date] = get_words_after(words, "תאריך סיום חלק זה של ההלוואה", 1)
        #get C: interest_type
        ws[interest_type] = get_words_after(words, "סוג הריבית", 1)
        #get D: current_interest
        ws[current_interest] = get_words_after(words, "שיעור הריבית המעודכנת להיום", 2) + get_words_after(words, "שיעור הריבית בחלק זה", 2)
        #get E: linkage_type
        ws[linkage_type] = get_words_after(words, "סוג ההצמדה", 2)
        #get F: balance
        ws[balance] = get_words_after(words, 'סיכום ביניים', 1)
        #get G: commission
        ws[commission] = get_words_after(words, 'סה"כ עמלת פרעון מוקדם', 1)
        #get H: loan_sum
        ws[loan_sum] = get_words_after(words, 'בחלק זה של ההלוואה', 2)
        #get I: payment_method
        ws[payment_method] = get_words_after(words, 'שיטת פרעון חלק זה בהלוואה', 2)

    # Save the file
    try:
        wb.save(xlsx)
        messagebox.showinfo("Success", "The data was extracted successfully.\nPlease check the output file, and adjust the parameters if needed.")
    except:
        messagebox.showinfo("Error", "The output file is open. Please close it and try again.")
        return
    #open the output file:
    import os
    os.system(f'start excel "{xlsx}"')



'''
this function grabs parameters from the GUI and calls the extract_data function
'''
def extract_data_from_gui():
    #if the pdf entry doesn't end with ".pdf", inform the user with a message box:
    if not pdf_entry.get().endswith(".pdf"):
        #inform the user that they need to choose a PDF file with a message box:
        messagebox.showinfo("Error", "You need to choose a PDF file")
        return
    #if the output entry doesn't end with ".xlsx", inform the user with a message box:
    if not output_entry.get().endswith(".xlsx"):
        #inform the user that they need to choose an output file with a message box:
        messagebox.showinfo("Error", "You need to choose an output file")
        return
    cells = []
    for i in range(len(fields)):
        key_entry = root.grid_slaves(row=i+4, column=1)[0]
        value_entry = root.grid_slaves(row=i+4, column=2)[0]
        cell_entry = root.grid_slaves(row=i+4, column=3)[0]
        cells.append(cell_entry.get())
        keys[fields[i]] = key_entry.get()
    try:
        extract_data(pdf_entry.get(), output_entry.get(), bank_var.get(), cells)
    except Exception as e:
        messagebox.showinfo("Error", e)
        return



'''
make a GUI for the user to choose the bank, the PDF file, and the output file, "key" strings, and the number of words to extract, and the cell to insert the data to
'''

try:
    root = tk.Tk()
    root.title('settings')

    output_label = tk.Label(root, text="Choose an output file:")
    output_label.grid(row=1, column=0)
    output_entry = tk.Entry(root)
    output_entry.grid(row=1, column=1)
    output_button = tk.Button(root, text="Choose output file", command=lambda: [output_entry.delete(0, tk.END), output_entry.insert(0, filedialog.asksaveasfilename())])
    output_button.grid(row=1, column=2)

    pdf_label = tk.Label(root, text="Choose a PDF file:")
    pdf_label.grid(row=0, column=0)
    pdf_entry = tk.Entry(root)
    pdf_entry.grid(row=0, column=1)
    pdf_button = tk.Button(root, text="Choose PDF file", command=lambda: [pdf_entry.delete(0, tk.END), pdf_entry.insert(0, filedialog.askopenfilename()), output_entry.delete(0, tk.END), output_entry.insert(0, pdf_entry.get().replace(".pdf", ".xlsx"))])
    pdf_button.grid(row=0, column=2)


    bank_label = tk.Label(root, text="Choose a bank:")
    bank_label.grid(row=2, column=0)
    bank_var = tk.StringVar(root)
    #list of the banks
    banks=['BenLeumi','Discount','Leumi', 'Poalim', 'Tfakhot']
    bank_var.set(banks[0])  # default value
    bank_menu = tk.OptionMenu(root, bank_var, *banks)
    bank_menu.grid(row=2, column=1)

    data_label = tk.Label(root, text="data")
    data_label.grid(row=3, column=0)
    keys_label = tk.Label(root, text="keys")
    keys_label.grid(row=3, column=1)
    values_label = tk.Label(root, text="values")
    values_label.grid(row=3, column=2)
    cells_label = tk.Label(root, text="cells")
    cells_label.grid(row=3, column=3)

    fields = ["report date", "end date", "interest type", "current interest", "linkage type", "balance", "commissions", "loan sum", "payment method"]
    for i, field in enumerate(fields):
        label = tk.Label(root, text=f"{field}:")
        label.grid(row=i+4, column=0)
        key_entry = tk.Entry(root)
        key_entry.grid(row=i+4, column=1)
        value_entry = tk.Entry(root)
        value_entry.grid(row=i+4, column=2)
        cell_entry = tk.Entry(root)
        cell_entry.grid(row=i+4, column=3)

    ok_button = tk.Button(root, text="OK", command=extract_data_from_gui)
    ok_button.grid(row=len(fields)+4, column=1, columnspan=2)

    #populate the entries with default values
    cells = ['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8', 'A9']
    for i, cell in enumerate(cells):
        cell_entry = root.grid_slaves(row=i+4, column=3)[0]
        cell_entry.insert(0, cell)
    #the keys are the strings to search for in the text, and are different for each bank
    keys = {
        'BenLeumi': ['', 'ת.סיום/סילוק משנה בפועל', 'סוג ריבית', 'ריבית נוכחית ליום מסירת המידע', 'סוג הצמדה', 'יתרה', 'עמלות', 'סכום ההלוואה', 'שיטת פירעון'],
        'Discount': ['תאריך הדפסה', 'מועד תשלום אחרון', 'סוג ריבית', 'שיעור ריבית שנתית', 'בסיס הצמדה', 'סה"כ יתרה', 'סה"כ עמלת פרעון מוקדם', 'סה"כ יתרה לסילוק', 'שיטת פרעון'],
        'Leumi': ['ליום:', 'תאריך סיום', 'סוג ריבית', 'ריבית שנתית להלוואה', 'בסיס הצמדה', 'יתרת קרן', 'סה"כ עמלות פירעון מוקדם', 'סה"כ יתרה לסילוק', 'שיטת פירעון ההלוואה'],
        'Poalim': ['תאריך :', 'מועד צפוי לתשלום', 'סוג ריבית', 'שיעור ריבית נומינלית', 'בסיס ההצמדה', 'עמלת פירעון מוקדם', 'יתרה לסילוק', 'שיטת פירעון ההלוואה'],
        'Tfakhot': ['יום עסקים', 'תאריך סיום חלק זה של ההלוואה', 'סוג הריבית', 'שיעור הריבית המעודכנת להיום', 'סוג ההצמדה', 'סיכום ביניים', 'סה"כ עמלת פרעון מוקדם', 'בחלק זה של ההלוואה', 'שיטת פרעון חלק זה בהלוואה']
    }

    root.mainloop()
except Exception as e:
    messagebox.showinfo("Error", e)
